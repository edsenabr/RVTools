#!/bin/bash
"exec" "$(dirname $0)/env/bin/python3" "$0" "$@"

import argparse
from datetime import datetime
from genericpath import exists
import sys
from price_loader import PriceList
from util import CellFormat
import openpyxl
import operator
from tqdm import tqdm

import os
import traceback

commit_colors = {
    "OD": "FFD8CE",
    "1Y":"FFF5CE",
    "3Y": "DDE8CB"
}

def get_parser(h):
    parser = argparse.ArgumentParser(add_help=h)
    parser.add_argument("-r", "--regions", nargs='*', help="region to be loaded", required=True)
    parser.add_argument("-p", "--period", nargs='?', help="regions to be loaded", default="monthly" , choices=['monthly', 'hourly'])
    parser.add_argument("-s", "--sheets", nargs='*', help="RVTools Spreadsheet", required=True)
    parser.add_argument("-nc", "--nocache", action='store_true', help="ignore cache")
    parser.add_argument("-l", "--local", action='store_true', help="use local html")
    return parser

def process_row(sheet, row_index, row, regions, columns):

        [vm, cpus, memory, disk, os_conf, os_tools] = operator.itemgetter(
            columns['VM'], 
            columns['CPUs'], 
            columns['Memory'], 
            columns['Unshared MB'], 
            columns['OS according to the configuration file'], 
            columns['OS according to the VMware Tools']
        )(row)
        disk = max(10, round(disk/1024, 2))

        os = os_conf or os_tools
        os_price = None
        if isinstance(os, str):
            os_price = price_list.get_os_price(os, cpus)


        color = 'EEEEEE' if not (row_index % 2) else None

        std = CellFormat(sheet).color(color)
        curr = CellFormat(sheet).color(color).currency()
        gb = CellFormat(sheet).color(color).gb()

        data=[
            std.value(vm), 
            std.value(cpus), 
            gb.value(memory/1024), 
            gb.value(disk), 
            std.value(os)
        ]

        for region in regions:
            od = price_list.select_price("od", cpus, memory, region)
            cud = price_list.select_price("cud1y", cpus, memory, region)
            data.extend([
                std.value( od["name"]),
                curr.value( od["od"]),
                std.value( cud["name"]),
                curr.value( cud["cud1y"]),
                std.value( cud["name"]),
                curr.value( cud["cud3y"]),
                curr.value( os_price),
                curr.value( price_list.select_disk_price("standard", region)*disk),
                curr.value( price_list.select_disk_price("balanced", region)*disk),
                curr.value( price_list.select_disk_price("ssd", region)*disk),
            ])
        sheet.append(data)
        return data

initial_row_offset=3
region_spacer=3

def get_gcve_offset(regions_qtty, books_qtty, header=False):
    region_rows = 3 + (3 if books_qtty == 1 else (3 * (books_qtty+1)))
    return initial_row_offset + (regions_qtty * (region_rows + region_spacer)) + (2 if not header else 0)

def write_book_header(sheet, regions, regions_qtty):
    sheet.append([
        CellFormat(sheet).header('DEE6EF').value('INPUT'),
        '','','','',
        CellFormat(sheet).header('E0C2CD').value('OUTPUT')
    ])

    row = [''] * 5
    for region_index, region in enumerate(regions):
        row.append(CellFormat(sheet).header(get_region_color(region_index)).value(region))
        row.extend([''] * 9)
    sheet.append(row)

    row = [''] * 5
    for region_index, region in enumerate(regions):
        row.append(CellFormat(sheet).header(get_region_color(region_index)).value('COMPUTE'))
        row.extend([''] * 5)
        row.append(CellFormat(sheet).header(get_region_color(region_index)).value('O.S.'))
        row.append(CellFormat(sheet).header(get_region_color(region_index)).value('DISK'))
        row.extend([''] * 2)
    sheet.append(row)

    row = [''] * 5
    for region in regions:
        row.append(CellFormat(sheet).header(commit_colors['OD']).value('ON DEMAND'))
        row.extend([''])
        row.append(CellFormat(sheet).header(commit_colors['1Y']).value('1 YEAR COMMIT'))
        row.extend([''])
        row.append(CellFormat(sheet).header(commit_colors['3Y']).value('3 YEARS COMMIT'))
        row.extend([''] * 5)
    sheet.append(row)


    row = ['VM', 'CPUS', 'Memory', 'Disk', 'OS']
    for region in regions:
        row.extend([
        *(['Family', 'Price']*3), 
        'License',
        'Standard', 
        'Balanced', 
        'SSD'
    ])
    sheet.append(CellFormat(sheet).header().generator(row))

    #input
    sheet.merge_cells(start_row=sheet.max_row-4, start_column=1, end_row=sheet.max_row-1, end_column=5)

    #output
    sheet.merge_cells(start_row=sheet.max_row-4, start_column=6, end_row=sheet.max_row-4, end_column=5 + (regions_qtty*10))

    for region_index, region in enumerate(regions):
        region_first_column = 6 + (region_index*10)
        region_last_column = 5 + ((region_index+1)*10)
        #name of the region
        sheet.merge_cells(start_row=sheet.max_row-3, start_column=region_first_column, end_row=sheet.max_row-3, end_column=region_last_column)

        #compute
        sheet.merge_cells(start_row=sheet.max_row-2, start_column=region_first_column, end_row=sheet.max_row-2, end_column=region_first_column+5)

        #os
        sheet.merge_cells(start_row=sheet.max_row-2, start_column=region_first_column+6, end_row=sheet.max_row-1, end_column=region_first_column+6)

        #disk
        sheet.merge_cells(start_row=sheet.max_row-2, start_column=region_first_column+7, end_row=sheet.max_row-1, end_column=region_last_column)

        #on demand
        sheet.merge_cells(start_row=sheet.max_row-1, start_column=region_first_column, end_row=sheet.max_row-1, end_column=region_first_column+1)

        #1 year commit
        sheet.merge_cells(start_row=sheet.max_row-1, start_column=region_first_column+2, end_row=sheet.max_row-1, end_column=region_first_column+3)

        #3 years commit
        sheet.merge_cells(start_row=sheet.max_row-1, start_column=region_first_column+4, end_row=sheet.max_row-1, end_column=region_first_column+5)


def is_currency(index):
    label = openpyxl.utils.get_column_letter(index)
    if index < 6:
        return False
    if (index -5) % 10 not in [1, 3, 5]:
        return True

    return False

def fit_sheet_columns(sheet):
    for index, column in enumerate(sheet.iter_cols(), 1):
        label = openpyxl.utils.get_column_letter(index)
        if index in [2,3,4] or is_currency(index) :
            sheet.column_dimensions[label].width= 10
        # elif is_currency(index):
        #     sheet.column_dimensions[label].width= 8
        else:
            length = max(len(str(cell.value))  for cell in column)
            sheet.column_dimensions[label].width = length

def fit_summary_columns(sheet):
    sheet.column_dimensions['A'].width=30
    sheet.column_dimensions['B'].width=8
    for colum_label in ['C', 'D', 'E', 'F']:
        sheet.column_dimensions[colum_label].width=15

def add_region_header(sheet, region_index, region_name):
    sheet.append(
        CellFormat(sheet)
            .header(get_region_color(region_index))
            .generator([
                region_name
            ])
    )
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=6)

    sheet.append(
        CellFormat(sheet)
            .header(get_region_color(region_index))
            .generator([
                'Input file',
                'Commit',
                'GCE w/ Disk Type',
                '','',
                'GCVE'
            ])
    )
    sheet.merge_cells(start_row=sheet.max_row, start_column=3, end_row=sheet.max_row, end_column=5)

    sheet.append(
        CellFormat(sheet)
            .header(get_region_color(region_index))
            .generator([
                '','',
                'SSD',
                'Balanced',
                'Standard',
                ''
            ])
    )
    sheet.merge_cells(start_row=sheet.max_row-1, start_column=1, end_row=sheet.max_row, end_column=1)
    sheet.merge_cells(start_row=sheet.max_row-1, start_column=2, end_row=sheet.max_row, end_column=2)
    sheet.merge_cells(start_row=sheet.max_row-1, start_column=6, end_row=sheet.max_row, end_column=6)


def get_region_color(region_index):
    return  ['DEE7E5', 'DEDCE6', 'F6F9D4'][region_index % 3]

def add_book_info(sheet, book_index, book_name, region_index, region_name, regions_qtty, books_qtty):
    gcve_price = price_list.get_gcve_price(region_name)
    if gcve_price is None:
        gcve_price = {"od": "NA()", "cud1y": "NA()", "cud3y":"NA()"}

    commits_name={
        1: "OD",
        3: "1Y",
        5: "3Y"
    }

    commits_column={
        1: "od",
        3: "cud1y",
        5: "cud3y"
    }

    start_colum = 6 + (region_index*10)
    gcve_offset = get_gcve_offset(regions_qtty, books_qtty) +  book_index

    header = CellFormat(sheet).header(get_region_color(region_index))
    first_row = sheet.max_row + 1
    for row_index, commit in enumerate([1,3,5]):
        format = CellFormat(sheet).color(commit_colors[commits_name[commit]]).currency()
        summary_row = [header.value(book_name), format.value(commits_name[commit])]
        for disk in [9,8,7]:
            summary_row.append(
            format.value("=SUM('{path}'!{commit}:{commit}, '{path}'!{os}:{os}, '{path}'!{disk}:{disk})".format(
                path=book_name,
                commit=openpyxl.utils.get_column_letter(start_colum+commit),
                os=openpyxl.utils.get_column_letter(start_colum+6),
                disk=openpyxl.utils.get_column_letter(start_colum+disk),
            )))
        summary_row.append(
            format.value('''=E{gcve_offset}*{price}'''.format(gcve_offset=gcve_offset, price=gcve_price[commits_column[commit]]))
        )
        sheet.append(summary_row)
    sheet.merge_cells(start_row=first_row, start_column=1, end_row=sheet.max_row, end_column=1)


def add_region_footer(sheet, books_qtty, region_index):
    if books_qtty > 1:

        last_sum_row = sheet.max_row
        first_sum_row = last_sum_row - 5
        header = CellFormat(sheet).header(get_region_color(region_index))

        for row_offset, commit in enumerate(['OD', '1Y', '3Y']):
            format = CellFormat(sheet).color(commit_colors[commit]).currency()
            summary_row = [header.value('TOTAL'), format.value(commit)]
            for column in ['C', 'D', 'E', 'F']:
                summary_row.append(
                    format.value('''=SUMPRODUCT((B{first}:B{last}="{commit}")*({column}{first}:{column}{last}))'''.format(first=first_sum_row, last=last_sum_row, commit=commit, column=column)),
                )
            sheet.append(summary_row)
            sheet.merge_cells(start_row=last_sum_row+1, start_column=1, end_row=last_sum_row+3, end_column=1)

    for i in range(0, region_spacer):
        sheet.append([])

def add_gcve_header(sheet):
    sheet.append(
        CellFormat(sheet)
            .header('FFDBB6')
            .generator([
                'GCVE'
            ])
    )
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=5)

    sheet.append(
        CellFormat(sheet)
            .header('FFDBB6')
            .generator(['Input file', 'vCPUS', 'Memory', 'Disk', 'Hosts'])
    )

def add_gcve_info(sheet, book_name):
    tb = CellFormat(sheet).tb()
    data = [
        CellFormat(sheet).center().value(book_name),
        "=SUM('{path}'!B:B)".format(path=book_name),
        tb.value("=SUM('{path}'!C:C)/1024".format(path=book_name)),
        tb.value("=SUM('{path}'!D:D)/1024".format(path=book_name)),
        '=ROUNDUP(max(B{row}/72,C{row}/768,D{row}/(19.2)))'.format(row=sheet.max_row + 1)
    ]
    sheet.append(data)

def add_gcve_footer(sheet, books_qtty):
    if books_qtty > 1:
        last_sum_row = sheet.max_row
        first_sum_row = last_sum_row - books_qtty
        bold = CellFormat(sheet).bold().color('FFDBB6')
        tb = CellFormat(sheet).bold().tb().color('FFDBB6')
        center = CellFormat(sheet).bold().color('FFDBB6').center()

        sheet.append([
            center.value('TOTAL'),
            bold.value('''=SUM(B{first}:B{last})'''.format(first=first_sum_row, last=last_sum_row)),
            tb.value('''=SUM(C{first}:C{last})'''.format(first=first_sum_row, last=last_sum_row)),
            tb.value('''=SUM(D{first}:D{last})'''.format(first=first_sum_row, last=last_sum_row)),
            bold.value('''=SUM(E{first}:E{last})'''.format(first=first_sum_row, last=last_sum_row))
        ])

def add_summary_disclamers(sheet, disclamers):
    for disclamer in disclamers:
        sheet.append([disclamer])
        sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=6)
    sheet.append([])
    initial_row_offset = len(disclamers) + 1

def process_file(book_name, output, regions, regions_qtty):
        book = openpyxl.load_workbook(book_name, read_only=True, data_only=True)
        columns = {}
        sheet = output.create_sheet(os.path.basename(book_name))
        write_book_header(sheet, regions, regions_qtty)
        input_sheet = book['vInfo']
        for row_index, row in enumerate(tqdm(input_sheet.iter_rows(values_only=True), desc=book_name, total=input_sheet.max_row)):
            if (row_index == 0): #header
                for index, column in enumerate(row):
                    columns[column] = index
            elif row[columns['VM']] is None:
                return
            else:
                try:
                    process_row(sheet, row_index, row, regions, columns)
                except Exception as e:
                    print("error processing row %s: %s" % (row_index, traceback.format_exc()))
        fit_sheet_columns(sheet)

def create_summary(summary, regions, regions_qtty, books, books_qtty):
    add_summary_disclamers(summary, ['*** on-demand prices includes sustained use discounts ***'])
    # add a summarization table per region to the Summary sheet
    for region_index, region_name in enumerate(regions):
        add_region_header(summary, region_index, region_name)
        for book_index, book_name in enumerate(books):
            add_book_info(summary, book_index, os.path.basename(book_name), region_index, region_name, regions_qtty, books_qtty)
        add_region_footer(summary, books_qtty, region_index)

    # add the gcve table to the Summary sheet
    add_gcve_header(summary)
    for book_name in books:
        add_gcve_info(summary, os.path.basename(book_name))
    add_gcve_footer(summary, books_qtty)
    fit_summary_columns(summary)    

def process_files(books, regions):
    books_qtty= len(books)
    regions_qtty= len(regions)
    output = openpyxl.Workbook()
    summary = output.active
    summary.title="Summary"

    # write one sheet per rvtools book to the target workbook
    for book_name in books:
        process_file(book_name, output, regions, regions_qtty)

    create_summary(summary, regions, regions_qtty, books, books_qtty)

    print("saving output file...")
    output.save(
        filename = 'estimated-rvtools-%s.xlsx' % datetime.now().strftime("%Y%m%d-%H%M%S")
    )
    print("...done.")
    output.close()

def validate_books(books):
    errors = []
    for book_name in books:
        book = openpyxl.load_workbook(book_name, read_only=True, data_only=True)
        if not 'vInfo' in book.sheetnames:
            errors.append("Sheet vInfo not found on %s" % book_name)
            continue
        # cols = list(book['vInfo'].iter_cols(min_row=1, max_row=1, values_only=True))
        required_cols = {
            'VM',
            'CPUs',
            'Memory',
            'Unshared MB',
            'OS according to the configuration file',
            'OS according to the VMware Tools'            
        }
        cols = [set(r) for r in book['vInfo'].iter_rows(min_row=1, max_row=1, values_only=True)][0]
        missing = required_cols - cols
        if len(missing) > 0:
            print("##################")
            print (missing)
            print("##################")
            sys.exit()

    if len(errors) > 0:
        print(*errors)
        sys.exit(127)


if (__name__=="__main__"):
    p = get_parser(h=True)
    args = p.parse_args()
    validate_books(args.sheets)
    price_list = PriceList(args.regions, args.period, args.nocache, args.local)
    process_files(args.sheets, args.regions)